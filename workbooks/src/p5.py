import datetime as dt, json
from openpyxl import load_workbook
from style import *
B="/tmp/claude-0/-home-user-bridgeapp-os/6418b099-52bd-5eb3-aab2-a68e43f06251/scratchpad/build/"
wb=load_workbook(B+"_p4.xlsx"); A=json.load(open(B+"addr.json")); RW=json.load(open(B+"rows.json"))
LOG="'05_Daily_Log'"; G1="'06_GE1_Paid_Media'"; G2="'07_GE2_School_Outreach'"; VIS="'08_School_Visits'"
F,LD=RW["FIRSTD"],RW["LASTD"]
START=dt.date(2026,8,24); DAYS=[START+dt.timedelta(days=i) for i in range(14)]

# ------------------------------------------------------- 08_School_Visits
ws=wb.create_sheet("08_School_Visits")
widths(ws,{"B":6,"C":30,"D":18,"E":14,"F":22,"G":11,"H":15,"I":16,"J":20,"K":14,"L":30})
title(ws,"SCHOOL VISITS","Owner: Schools Coordinator. Logged the same day as the visit. Week 1 is assigned from the six active Phase 2 pilots. Week 2 is assigned at the day 8 stand up.")
r=5
r=band(ws,r,"Week 1: Tue 25, Wed 26, Thu 27 August",10)
r=header(ws,r,["#","School","Region","Date","Owner","Status","Parents reached","Signups captured","UTM content tag","Sales attributed","Notes"])
VF=r
W1=[("Kingsmead","KZN (Durban)",DAYS[1]),("St Albans","EC (Port Elizabeth)",DAYS[1]),
    ("Dainfern College","GP (JHB North)",DAYS[2]),("Umtata International","EC (Mthatha)",DAYS[2]),
    ("St Jude Private","EC",DAYS[3]),("Christ The King International","EC",DAYS[3])]
i=1
for name,reg,d in W1:
    ws.cell(row=r,column=2,value=i).font=BODY_B; ws.cell(row=r,column=2).alignment=CTR
    c=ws.cell(row=r,column=3,value=name); c.font=INPUT; c.fill=F_YELLOW
    ws.cell(row=r,column=4,value=reg).font=MUTED
    c=ws.cell(row=r,column=5,value=d); c.number_format=DAT; c.font=BODY_B
    c=ws.cell(row=r,column=6,value="TBC"); c.font=INPUT; c.fill=F_YELLOW
    c=ws.cell(row=r,column=7,value="PLANNED"); c.font=INPUT; c.fill=F_YELLOW; c.alignment=CTR
    for col in (8,9,11):
        c=ws.cell(row=r,column=col); c.font=INPUT; c.fill=F_YELLOW; c.number_format=NUM; c.alignment=RGT
    c=ws.cell(row=r,column=10,value="school-%s"%name.lower().split()[0]); c.font=INPUT; c.fill=F_YELLOW
    c=ws.cell(row=r,column=12); c.font=INPUT; c.fill=F_YELLOW
    r+=1; i+=1
r+=1
r=band(ws,r,"Week 2: Tue 1, Wed 2, Thu 3 September. Schools assigned at the day 8 stand up",10)
r=header(ws,r,["#","School","Region","Date","Owner","Status","Parents reached","Signups captured","UTM content tag","Sales attributed","Notes"])
for d in [DAYS[8],DAYS[8],DAYS[9],DAYS[9],DAYS[10],DAYS[10]]:
    ws.cell(row=r,column=2,value=i).font=BODY_B; ws.cell(row=r,column=2).alignment=CTR
    for col in (3,4,6,7,8,9,10,11,12):
        c=ws.cell(row=r,column=col); c.font=INPUT; c.fill=F_YELLOW
        if col in (8,9,11): c.number_format=NUM; c.alignment=RGT
    ws.cell(row=r,column=7,value="PLANNED").alignment=CTR
    c=ws.cell(row=r,column=5,value=d); c.number_format=DAT; c.font=BODY_B
    r+=1; i+=1
VL=r-1
r+=1
ws.cell(row=r,column=2,value="Reference: the Phase 2 pool is 6 active pilots (Kingsmead, St Albans, Dainfern College, Umtata International, St Jude Private, Christ The King International) plus 2 backups (Strategic High, Excelsior High). Source: v4 sheet 07_Pilot_Schools.").font=MUTED_I
r+=1
ws.cell(row=r,column=2,value="Every visit carries a QR code whose link is registered in 09_UTM_Register with utm_content set to the tag in column J. A visit with no tag produces sales nobody can attribute to it.").font=MUTED_I
r+=1
ws.cell(row=r,column=2,value="Signups captured here must equal the GE2 signups captured entered in 05_Daily_Log for the same date.").font=MUTED_I

# ------------------------------------------------------- 09_UTM_Register
LIVE_UTM = "https://docs.google.com/spreadsheets/d/1D7rVlQW8ym0bb20i3-7qmEHBS5fxVVRegwjdhaK9JF8"
ws=wb.create_sheet("09_UTM_Register")
widths(ws,{"B":34,"C":16,"D":14,"E":20,"F":24,"G":7,"H":12,"I":18,"J":12})
title(ws,"UTM CODES",
 "There is one live register and it is not this tab. Links are built and stored in the BridgeApp UTM link builder in Google Drive, so the whole team works off one list.")
r=5
r=band(ws,r,"Where links live",8)
for a,b in [
 ("Live register","BridgeApp UTM link builder, Google Sheets. Every link is created there and nowhere else."),
 ("Link", LIVE_UTM),
 ("Why not here","A workbook file gets emailed and forked. A Sheet does not. Two registers means links tagged in one and counted from the other."),
 ("This tab","The checklist of codes this sprint needs. Tick each one off as it is created in the register."),
]:
    ws.cell(row=r,column=2,value=a).font=BODY_B
    c=ws.cell(row=r,column=3,value=b); c.font=BODY if a!="Link" else Font(name="Arial",sz=9,color=BLUE)
    c.alignment=WRAP
    ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=10)
    ws.row_dimensions[r].height=24
    r+=1
r+=1
r=band(ws,r,"Codes this sprint needs. Tick column H when it exists in the live register.",8)
r=header(ws,r,["Purpose","Source","Medium","Campaign","Content","GE","Created? Y/N","Owner","Live from"])
UF=r
ROWS=[
 ("Meta paid, parent audience","meta","paid-social","qualifying-sprint","meta-set-a","GE1"),
 ("Meta paid, retargeting July leads","meta","paid-social","qualifying-sprint","meta-retarget","GE1"),
 ("Meta paid, 20 second recut","meta","paid-social","qualifying-sprint","recut-20s","GE1"),
 ("Meta paid, static statement","meta","paid-social","qualifying-sprint","static-statement","GE1"),
 ("Boosted organic, best of last week","facebook","boosted","qualifying-sprint","weekly-boost","GE1"),
 ("Instagram bio link","instagram","organic-social","qualifying-sprint","bio-link","GE1"),
 ("Nurture emailer","email","email","qualifying-sprint","nurture-1","GE1"),
 ("School QR, Pretoria Technical","school-visits","qr","qualifying-sprint","school-pretoria-technical","GE2"),
 ("School QR, Pretoria-Wes","school-visits","qr","qualifying-sprint","school-pretoria-wes","GE2"),
 ("School QR, Pretoria Girls","school-visits","qr","qualifying-sprint","school-pretoria-girls","GE2"),
 ("School QR, generic fallback","school-visits","qr","qualifying-sprint","","GE2"),
 ("Learner leave behind, printed","school-visits","print","qualifying-sprint","leave-behind","GE2"),
 ("Parent take home slip","school-visits","print","qualifying-sprint","parent-slip","GE2"),
]
for purpose,src,med,camp,cont,ge in ROWS:
    ws.cell(row=r,column=2,value=purpose).font=BODY
    for col,v in [(3,src),(4,med),(5,camp),(6,cont)]:
        c=ws.cell(row=r,column=col,value=v); c.font=BODY; c.fill=F_GREY
    g=ws.cell(row=r,column=7,value=ge); g.font=BODY_B; g.alignment=CTR
    g.fill=F_BLUE if ge=="GE1" else F_ORANGE
    c=ws.cell(row=r,column=8); c.font=INPUT; c.fill=F_YELLOW; c.alignment=CTR
    c=ws.cell(row=r,column=9,value="TBC"); c.font=INPUT; c.fill=F_YELLOW
    c=ws.cell(row=r,column=10); c.font=INPUT; c.fill=F_YELLOW; c.number_format=DAT
    r+=1
UL=r-1
r+=1
r=band(ws,r,"Rules",8)
for t in ["Source and medium are compulsory. Campaign and content are optional.",
          "Everything is lower case with spaces as hyphens, because GA4 treats Meta and meta as two different sources.",
          "The landing URL is the parent page. If it is not deployed, repoint every link in the register before any spend starts.",
          "A school visited without its own code produces sales nobody can trace back to the room it started in.",
          "A link that is not in the live register cannot be attributed, so the sale it produces counts for nobody."]:
    ws.cell(row=r,column=2,value=t).font=MUTED_I
    ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=10)
    r+=1
json.dump({"VF":VF,"VL":VL,"UF":UF,"UL":UL},open(B+"rows2.json","w"))
wb.save(B+"_p5.xlsx"); print("p5 ok",VF,VL,UF,UL)
