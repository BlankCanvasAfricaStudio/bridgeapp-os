import json
from openpyxl import load_workbook
from style import *
B="/tmp/claude-0/-home-user-bridgeapp-os/6418b099-52bd-5eb3-aab2-a68e43f06251/scratchpad/build/"
wb=load_workbook(B+"_p1.xlsx")
BAS="'02_Baselines'"; K={}

ws=wb.create_sheet("02_Baselines")
widths(ws,{"B":42,"C":16,"D":13,"E":58})
title(ws,"BASELINES: WHAT THE PLATFORMS ALREADY TELL US",
 "Real reported figures, not estimates. Every driver on 03_Sprint_Drivers that can be grounded in one of these is grounded in it.")
r=5
def row(r,lab,val,fmt,src,key=None,calc=False):
    ws.cell(row=r,column=2,value=lab).font=BODY
    c=ws.cell(row=r,column=3,value=val); c.number_format=fmt; c.alignment=RGT
    c.font=BODY_B if calc else INPUT
    if calc: c.fill=F_GREY
    ws.cell(row=r,column=5,value=src).font=MUTED_I
    if key: K[key]="%s!$C$%d"%(BAS,r)
    return r+1

r=band(ws,r,"Paid: Meta boosted post, 29 July to 5 August 2026, 8 days",4)
r=header(ws,r,["Reported figure","Value","","Source"])
SRC="Meta Ad insights, boosted Instagram post, screenshot 20 Aug 2026."
r=row(r,"Spend",1617.57,CUR2,SRC,"spend")
r=row(r,"Daily budget",232.00,CUR2,SRC)
r=row(r,"Views",15979,NUM,SRC)
r=row(r,"Viewers reached",11034,NUM,SRC)
r=row(r,"Link clicks",491,NUM,SRC,"clicks")
r=row(r,"Link click through rate",0.031,PCT,SRC)
r=row(r,"Cost per link click","=%s/%s"%(K["spend"],K["clicks"]),CUR2,
      "Calculated from the two rows above. This is the only cost figure in the campaign that is real.","cpc",calc=True)
r=row(r,"Instagram profile visits",490,NUM,SRC)
r=row(r,"Cost per profile visit",3.30,CUR2,SRC)
r=row(r,"Instagram follows",127,NUM,SRC)
r=row(r,"Post engagements",6595,NUM,SRC)
r=row(r,"Reactions",417,NUM,SRC)
r=row(r,"Comments",1,NUM,SRC)
r=row(r,"Shares",30,NUM,SRC)
r=row(r,"Saves",52,NUM,SRC)
r=row(r,"Reel plays",14730,NUM,SRC)
r=row(r,"ThruPlays",1757,NUM,SRC)
r=row(r,"Video length (seconds)",53,NUM,SRC)
r=row(r,"Average watch time (seconds)",7,NUM,SRC)
r=row(r,"Hook rate",0.380,PCT,SRC)
r=row(r,"Hold rate",0.076,PCT,SRC)
r+=1
r=band(ws,r,"Paid: who it actually reached",4)
r=header(ws,r,["Reported figure","Value","","Source"])
r=row(r,"Women",0.607,PCT,SRC)
r=row(r,"Men",0.384,PCT,SRC)
r=row(r,"Non binary",0.009,PCT,SRC)
r=row(r,"Largest age band","35 to 44",'General',"Meta reports bands as bars without figures. 35 to 44 is the longest bar, 25 to 34 second.")
r=row(r,"Gauteng","Largest",'General',"Bar is full width, figure not shown in the export.")
r=row(r,"KwaZulu Natal",1200,NUM,SRC)
r=row(r,"Western Cape",863,NUM,SRC)
r=row(r,"Limpopo",572,NUM,SRC)
r=row(r,"Eastern Cape",556,NUM,SRC)
r=row(r,"Mpumalanga",556,NUM,SRC)
r+=1
r=band(ws,r,"Organic: reel on Facebook and Instagram",4)
r=header(ws,r,["Reported figure","Value","","Source"])
S2="Reel insights, screenshot 20 Aug 2026."
r=row(r,"Views, total",12677,NUM,S2)
r=row(r,"Views, Instagram",11992,NUM,S2)
r=row(r,"Views, Facebook",685,NUM,S2)
r=row(r,"Views from followers",1410,NUM,S2)
r=row(r,"Views from non followers",10977,NUM,S2)
r=row(r,"Interactions",69,NUM,S2)
r=row(r,"Reactions",55,NUM,S2)
r=row(r,"Shares",9,NUM,S2)
r=row(r,"Saves",5,NUM,S2)
r=row(r,"Follows",17,NUM,S2)
r+=1
r=band(ws,r,"Organic: TikTok, single video posted 13 March 2026",4)
r=header(ws,r,["Reported figure","Value","","Source"])
S3="TikTok Studio video analysis, screenshot 20 Aug 2026."
r=row(r,"Video views",38922,NUM,S3)
r=row(r,"Total viewers",30800,NUM,S3)
r=row(r,"Average watch time (seconds)",16.9,'0.0',S3)
r=row(r,"Watched full video",0.0625,PCT,S3)
r=row(r,"Likes",5194,NUM,S3)
r=row(r,"Comments",20,NUM,S3)
r=row(r,"Shares",37,NUM,S3)
r=row(r,"Saves",159,NUM,S3)
r=row(r,"New followers",25,NUM,S3)
r=row(r,"Returning viewers",0.97,PCT,S3)
r=row(r,"Non followers",0.60,PCT,S3)
r=row(r,"Female",0.65,PCT,S3)
r=row(r,"Male",0.35,PCT,S3)
r=row(r,"Age 18 to 24",0.54,PCT,S3)
r=row(r,"Age 25 to 34",0.33,PCT,S3)
r=row(r,"Age 35 to 44",0.08,PCT,S3)
r=row(r,"South Africa",0.779,PCT,S3)
r+=1
r=band(ws,r,"What this settles, and what it does not",4)
for a,b in [
 ("Settled","Cost per link click on Meta, at R3.29. That is a real number from a real buy against a real audience."),
 ("Settled","Meta reaches the paying adult. Women 60.7%, largest band 35 to 44. That is the parent."),
 ("Settled","TikTok reaches the student, not the buyer. 54% aged 18 to 24, 97% returning viewers. Treat it as demand creation, not as a sales channel."),
 ("Not settled","Click to signup. The lead endpoint is not live, so nobody knows how many of the 491 clicks became a signup."),
 ("Not settled","Signup to sale. GA4 has no purchase event, so no click has ever been tied to a payment."),
 ("Not settled","Therefore cost per sale. It cannot be derived from anything held today. Measuring it is the entire purpose of this sprint."),
]:
    c=ws.cell(row=r,column=2,value=a); c.font=GOOD if a=="Settled" else WARN
    c=ws.cell(row=r,column=3,value=b); c.font=BODY; c.alignment=WRAP
    ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=5)
    ws.row_dimensions[r].height=26
    r+=1
json.dump(K,open(B+"bas.json","w"))
wb.save(B+"_p1b.xlsx"); print("p1b ok"); print(json.dumps(K,indent=0))
