from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

NAVY="FF0F2A4A"; BLUE="FF4A90D9"; ORANGE="FFF97316"; YELLOW="FFFBBF24"
SLATE="FF475569"; INK="FF0B1220"; WHITE="FFFFFFFF"; INPUT_BLUE="FF0000FF"
GREEN="FF15803D"; RED="FFB91C1C"

F_NAVY   = PatternFill("solid", fgColor=NAVY)
F_YELLOW = PatternFill("solid", fgColor="FFFFFF00")
F_CREAM  = PatternFill("solid", fgColor="FFFEF7E7")
F_BLUE   = PatternFill("solid", fgColor="FFE6F0FB")
F_ORANGE = PatternFill("solid", fgColor="FFFEE9D6")
F_GREY   = PatternFill("solid", fgColor="FFEEF2F7")

TITLE   = Font(name="Arial", sz=20, b=True, color=NAVY)
SUB     = Font(name="Arial", sz=9,  i=True, color=SLATE)
H2      = Font(name="Arial", sz=14, b=True, color=NAVY)
BAND    = Font(name="Arial", sz=11, b=True, color=WHITE)
HEAD    = Font(name="Arial", sz=10, b=True, color=WHITE)
BODY    = Font(name="Arial", sz=10, color=INK)
BODY_B  = Font(name="Arial", sz=10, b=True, color=INK)
MUTED   = Font(name="Arial", sz=10, color=SLATE)
MUTED_I = Font(name="Arial", sz=9,  i=True, color=SLATE)
ACCENT  = Font(name="Arial", sz=10, b=True, color=BLUE)
WARN    = Font(name="Arial", sz=10, b=True, color=ORANGE)
INPUT   = Font(name="Arial", sz=10, b=True, color=INPUT_BLUE)
NAVY_B  = Font(name="Arial", sz=10, b=True, color=NAVY)
GOOD    = Font(name="Arial", sz=10, b=True, color=GREEN)
BAD     = Font(name="Arial", sz=10, b=True, color=RED)

CUR = '"R"#,##0;("R"#,##0);\\-'
NUM = '#,##0;(#,##0);\\-'
PCT = '0.0%'
DAT = 'ddd dd mmm'
CUR2= '"R"#,##0.00;("R"#,##0.00);\\-'

WRAP = Alignment(wrap_text=True, vertical="top")
TOP  = Alignment(vertical="top")
CTR  = Alignment(horizontal="center", vertical="center")
RGT  = Alignment(horizontal="right", vertical="center")

_thin = Side(style="thin", color="FFD8DEE7")
BOX = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
UNDER = Border(bottom=_thin)

def band(ws, row, text, ncols, first=2):
    """Full-width navy section band."""
    c = ws.cell(row=row, column=first, value=text)
    c.font = BAND; c.fill = F_NAVY; c.alignment = Alignment(vertical="center")
    for i in range(first, first+ncols):
        ws.cell(row=row, column=i).fill = F_NAVY
    ws.row_dimensions[row].height = 20
    return row+1

def header(ws, row, labels, first=2, fill=None):
    """Column header row."""
    fill = fill or PatternFill("solid", fgColor="FF1E3A5F")
    for i, lab in enumerate(labels):
        c = ws.cell(row=row, column=first+i, value=lab)
        c.font = HEAD; c.fill = fill; c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[row].height = 26
    return row+1

def title(ws, t, sub, ncols=8):
    ws.cell(row=2, column=2, value=t).font = TITLE
    ws.row_dimensions[2].height = 27
    ws.cell(row=3, column=2, value=sub).font = SUB
    ws.row_dimensions[3].height = 13

def widths(ws, spec):
    from openpyxl.utils import get_column_letter
    ws.column_dimensions["A"].width = 3
    for col, w in spec.items():
        ws.column_dimensions[col].width = w

def note(ws, row, text, first=2):
    c = ws.cell(row=row, column=first, value=text); c.font = MUTED_I
    return row+1
