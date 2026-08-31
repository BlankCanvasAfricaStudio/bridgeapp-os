import datetime as dt
from openpyxl import Workbook
from openpyxl.utils import get_column_letter as L
from style import *

DRV = "'03_Sprint_Drivers'"
LOG = "'05_Daily_Log'"
START = dt.date(2026, 8, 24)
DAYS = [START + dt.timedelta(days=i) for i in range(14)]

wb = Workbook()

# ----------------------------------------------------------------- 00_README
ws = wb.active; ws.title = "00_README"
widths(ws, {"B": 24, "C": 92})
title(ws, "BRIDGEAPP: QUALIFYING SPRINT",
      "14 days, Mon 24 Aug to Sun 6 Sep 2026. Purpose: establish the real cost per sale on both growth engines before committing the full budget.")
r = 5
r = band(ws, r, "How this workbook is used", 2)
for a, b in [
    ("Every morning",  "The named owner fills yesterday's row in 05_Daily_Log. Yellow cells are the only cells anyone types into."),
    ("Every Monday",   "10_KPIs is read at stand-up. Any KPI on its trigger is discussed before anything else."),
    ("Day 14 (6 Sep)", "11_Go_No_Go is completed. That sheet decides whether the full budget is released, held, or re-cut."),
    ("Never",          "Do not type over a black cell. Black cells are formulas and will stop calculating."),
]:
    ws.cell(row=r, column=2, value=a).font = BODY_B
    ws.cell(row=r, column=3, value=b).font = BODY
    r += 1
r += 1
r = band(ws, r, "Sheet guide", 2)
for a, b in [
    ("00_README", "This sheet."),
    ("01_Owners", "Named owner per sheet and per decision. Fill the Named Person column before day 1."),
    ("02_Baselines", "Real reported figures from Meta, Instagram, Facebook and TikTok. The only hard numbers the campaign holds."),
    ("03_Sprint_Drivers", "Every assumption in one place. Change a yellow cell here and the whole workbook recalculates."),
    ("04_Sprint_Plan", "The 14 days, day by day. GE tag, deliverable, owner, due, status."),
    ("05_Daily_Log", "The data entry sheet. Everything else in this workbook reads from it."),
    ("06_GE1_Paid_Media", "Paid media engine. Spend, clicks, signups, sales, cost per sale. Reads 05_Daily_Log."),
    ("07_GE2_School_Outreach", "School outreach engine. Visits, calls, signups captured, sales. Reads 05_Daily_Log."),
    ("08_School_Visits", "The visit schedule for these 14 days, and the result of each visit."),
    ("09_UTM_Register", "Every tagged link used in the sprint. A link that is not on this sheet does not get counted."),
    ("10_KPIs", "Live dashboard by engine. Current pulls from 05_Daily_Log. Target and Trigger are fixed."),
    ("11_Go_No_Go", "The day 14 decision gate, with the thresholds written before the sprint starts."),
]:
    ws.cell(row=r, column=2, value=a).font = ACCENT
    ws.cell(row=r, column=3, value=b).font = BODY
    r += 1
r += 1
r = band(ws, r, "Colour legend", 2)
for f, a, b in [
    (F_YELLOW, "Yellow fill",  "An input. Type here. These are the only cells that should ever be edited."),
    (None,     "Blue bold text","A number that was typed in, not calculated. Change it in 03_Sprint_Drivers only."),
    (None,     "Black text",   "A formula. Do not overwrite."),
    (F_ORANGE, "Orange fill",  "Blocked on someone outside this team. Named in the row."),
]:
    c = ws.cell(row=r, column=2, value=a)
    c.font = INPUT if b.startswith("A number") else BODY_B
    if f: c.fill = f
    ws.cell(row=r, column=3, value=b).font = BODY
    r += 1
r += 1
r = note(ws, r, "Source for all pre-filled targets and budget lines: BridgeApp Phase 2 Operating Workbook v4 (May 2026), sheets 02_Budget_Master, 03_Calendar_Phase2, 04_GE1_Tracker, 05_GE2_Tracker, 09_KPIs_By_GE.")
r = note(ws, r, "Source for the 40 000 by 30 Nov 2026 target and the 15 week pace: campaign brief, Blank Canvas.")
r = note(ws, r, "Counting sales is a Gradesmatch tech function, not a manual task. It needs either the GA4 purchase event or an automated feed of cleared Peach transactions. Until one of them lands, this workbook has no way to record a sale. See 03_Sprint_Drivers.")

# ----------------------------------------------------------------- 01_Owners
ws = wb.create_sheet("01_Owners")
widths(ws, {"B": 34, "C": 24, "D": 22, "E": 34})
title(ws, "OWNERS", "One name per sheet. Fill the yellow Named Person column before day 1. An unowned sheet does not get filled in.")
r = 5
r = band(ws, r, "Sheet ownership", 4)
r = header(ws, r, ["Sheet / workstream", "Owner role", "Named person", "Cadence"])
own = [
    ("03_Sprint_Drivers", "Campaign Lead", "Weekly, Monday"),
    ("04_Sprint_Plan", "Campaign Lead", "Daily"),
    ("05_Daily_Log (GE1 columns)", "Performance Marketer", "Daily by 09:00"),
    ("05_Daily_Log (GE2 columns)", "Schools Coordinator", "Same day as visit"),
    ("06_GE1_Paid_Media", "Performance Marketer", "Daily"),
    ("07_GE2_School_Outreach", "Schools Coordinator", "Daily"),
    ("08_School_Visits", "Schools Coordinator", "Same day as visit"),
    ("09_UTM_Register", "Performance Marketer", "Before any link goes live"),
    ("10_KPIs", "Campaign Lead", "Weekly, Monday stand-up"),
    ("11_Go_No_Go", "Campaign Lead", "Once, day 14"),
    ("Sales counting, Peach to ledger", "Gradesmatch tech", "Automated, daily"),
    ("Creative for paid + school assets", "Designer", "Per brief"),
    ("Organic content in support of sprint", "Content Producer", "Daily"),
]
for a, b, c in own:
    ws.cell(row=r, column=2, value=a).font = BODY
    ws.cell(row=r, column=3, value=b).font = BODY
    cc = ws.cell(row=r, column=4, value="TBC"); cc.font = INPUT; cc.fill = F_YELLOW
    ws.cell(row=r, column=5, value=c).font = MUTED
    r += 1
r += 1
r = band(ws, r, "Decision authority for these 14 days", 4)
r = header(ws, r, ["Decision", "Approver", "Escalation", "Notes"])
dec = [
    ("Move spend between GE1 and GE2", "Campaign Lead", "MD", "Inside the sprint cap only"),
    ("Any spend above the sprint cap", "MD", "Founders", "Pre approval, in writing"),
    ("Pause a channel mid sprint", "Performance Marketer", "Campaign Lead", "Same day if cost per sale doubles target"),
    ("Add or drop a school visit", "Schools Coordinator", "Campaign Lead", "Log the reason in 08_School_Visits"),
    ("Publish a link without a UTM", "Not permitted", "Campaign Lead", "Untagged traffic cannot be counted"),
    ("Change a driver in 03_Sprint_Drivers", "Campaign Lead", "MD", "Note the date and reason in the row"),
    ("Day 14 go or no go", "MD", "Founders", "On the evidence in 11_Go_No_Go"),
]
for a, b, c, d in dec:
    ws.cell(row=r, column=2, value=a).font = BODY
    ws.cell(row=r, column=3, value=b).font = BODY_B
    ws.cell(row=r, column=4, value=c).font = BODY
    ws.cell(row=r, column=5, value=d).font = MUTED
    r += 1

wb.save("/tmp/claude-0/-home-user-bridgeapp-os/6418b099-52bd-5eb3-aab2-a68e43f06251/scratchpad/build/_p1.xlsx")
print("p1 ok")
