import datetime as dt, json
from openpyxl import load_workbook
from style import *
B="/tmp/claude-0/-home-user-bridgeapp-os/6418b099-52bd-5eb3-aab2-a68e43f06251/scratchpad/build/"
wb=load_workbook(B+"_p2.xlsx"); A=json.load(open(B+"addr.json"))
START=dt.date(2026,8,24); DAYS=[START+dt.timedelta(days=i) for i in range(14)]

# --------------------------------------------------------- 04_Sprint_Plan
ws=wb.create_sheet("04_Sprint_Plan")
widths(ws,{"B":6,"C":14,"D":9,"E":50,"F":22,"G":13,"H":12,"I":16})
title(ws,"SPRINT PLAN: 14 DAYS",
 "Every commitment in the window. GE3 rows are carried from the Phase 2 calendar; they consume team time but carry no sales target in this sprint.")
r=5
r=band(ws,r,"Mon 24 Aug to Sun 6 Sep 2026",8)
r=header(ws,r,["Day","Date","GE","Deliverable","Owner","Due","Status","Source"])
FIRST=r
PLAN=[
(1,"ALL","Sprint drivers locked and signed off","Campaign Lead","24 Aug","Sprint"),
(1,"ALL","Every role in 01_Owners has a named person","Campaign Lead","24 Aug","Sprint"),
(1,"GE1","09_UTM_Register complete. No link goes live untagged","Performance Marketer","24 Aug","Sprint"),
(1,"ALL","Baseline: closing sales count recorded in 05_Daily_Log","Campaign Lead","24 Aug","Blocked: dev"),
(1,"GE1","Paid buy live on Meta against the sprint cap","Performance Marketer","24 Aug","Sprint"),
(1,"GE2","Week 1 visit schedule confirmed with each school","Schools Coordinator","24 Aug","Sprint"),
(1,"ALL","Escalation sent to Gradesmatch dev on the three blocked items","Campaign Lead","24 Aug","Blocked: dev"),
(2,"GE1","First 24 hour read. Pull any creative below the CTR floor","Performance Marketer","25 Aug","Sprint"),
(2,"GE2","School visits 1 and 2","Schools Coordinator","25 Aug","Sprint"),
(2,"GE1","FC 25K Push continues, paid and organic","Performance Marketer","Continuous","v4 Calendar"),
(3,"GE2","School visits 3 and 4","Schools Coordinator","26 Aug","Sprint"),
(3,"GE1","Creative variant set 2 live","Designer","26 Aug","Sprint"),
(3,"GE3","BMW Golf Clinic invitee list sent","Strategy Lead","26 Aug","v4 Calendar"),
(4,"GE1","UniApply matric urgency push live","Performance Marketer","27 Aug","v4 Calendar"),
(4,"GE2","School visits 5 and 6","Schools Coordinator","27 Aug","Sprint"),
(4,"ALL","Mid week check that the Peach feed is landing","Campaign Lead","27 Aug","Sprint"),
(5,"GE1","Influencer Wave 3, GP creator, content live","Creator Manager","28 Aug","v4 Calendar"),
(5,"GE2","Week 1 visit results logged in 08_School_Visits","Schools Coordinator","28 Aug","Sprint"),
(5,"ALL","Week 1 working book updated","Campaign Lead","28 Aug","Sprint"),
(5,"GE3","Podcast pilot episode format lock","Content Producer","28 Aug","v4 Calendar"),
(6,"GE1","Weekend spend read. No new creative","Performance Marketer","29 Aug","Sprint"),
(6,"GE1","Organic conversion post, both channels","Content Producer","29 Aug","Sprint"),
(7,"GE1","Week 1 close. Cost per sale for week 1 computed","Performance Marketer","30 Aug","Sprint"),
(8,"ALL","Week 1 review at stand up. 10_KPIs read before anything else","Campaign Lead","31 Aug","Sprint"),
(8,"GE1","Reallocate spend across channels, inside the cap","Performance Marketer","31 Aug","Sprint"),
(8,"GE2","Week 2 visit schedule confirmed","Schools Coordinator","31 Aug","Sprint"),
(8,"GE3","BMW and Porsche negotiation close, signature","Strategy Lead","31 Aug","v4 Calendar"),
(9,"GE2","School visits 7 and 8","Schools Coordinator","1 Sep","Sprint"),
(9,"GE2","Conference follow up and MOU push begins","Schools Coordinator","1 Sep","v4 Calendar"),
(9,"GE2","SAHISA Durban prep begins, conference 13 to 16 Sep","Schools Coordinator","1 Sep","v4 Calendar"),
(9,"GE1","Creative variant set 3 live","Designer","1 Sep","Sprint"),
(10,"GE2","School visits 9 and 10","Schools Coordinator","2 Sep","Sprint"),
(10,"GE1","Mid sprint cost per sale read against target","Performance Marketer","2 Sep","Sprint"),
(11,"GE2","School visits 11 and 12","Schools Coordinator","3 Sep","Sprint"),
(11,"GE1","Pause any channel running above twice target cost per sale","Performance Marketer","3 Sep","Sprint"),
(12,"ALL","Week 2 working book updated","Campaign Lead","4 Sep","Sprint"),
(12,"GE2","Week 2 visit results logged","Schools Coordinator","4 Sep","Sprint"),
(13,"GE1","Organic conversion post, both channels","Content Producer","5 Sep","Sprint"),
(13,"GE1","Weekend spend read","Performance Marketer","5 Sep","Sprint"),
(14,"ALL","Final sales count confirmed against the Peach feed","Campaign Lead","6 Sep","Blocked: dev"),
(14,"ALL","11_Go_No_Go completed and sent to MD","Campaign Lead","6 Sep","Sprint"),
]
GEF={"GE1":F_BLUE,"GE2":F_ORANGE,"GE3":F_GREY,"ALL":F_CREAM}
prev=None
for d,ge,deliv,own,due,src in PLAN:
    ws.cell(row=r,column=2,value=d).font=BODY_B; ws.cell(row=r,column=2).alignment=CTR
    if d!=prev:
        c=ws.cell(row=r,column=3,value=DAYS[d-1]); c.number_format=DAT; c.font=BODY_B
    prev=d
    g=ws.cell(row=r,column=4,value=ge); g.font=BODY_B; g.fill=GEF[ge]; g.alignment=CTR
    ws.cell(row=r,column=5,value=deliv).font=BODY
    ws.cell(row=r,column=6,value=own).font=BODY
    ws.cell(row=r,column=7,value=due).font=MUTED
    s=ws.cell(row=r,column=8,value="PLANNED"); s.font=INPUT; s.fill=F_YELLOW; s.alignment=CTR
    sc=ws.cell(row=r,column=9,value=src)
    sc.font=WARN if src.startswith("Blocked") else MUTED_I
    r+=1
LAST=r-1
r+=1
ws.cell(row=r,column=2,value="Status values: PLANNED / DOING / DONE / SLIPPED / DROPPED. Anything not DONE on its due date is raised at the next stand up.").font=MUTED_I
r+=1
ws.cell(row=r,column=2,value="School visits: 6 per week is the visit target. The Phase 2 v4 GE2 KPI of 25 per week is calls plus visits combined, so 19 calls per week sit alongside these.").font=MUTED_I

# ----------------------------------------------------------- 05_Daily_Log
ws=wb.create_sheet("05_Daily_Log")
widths(ws,{"B":14,"C":6,"D":13,"E":11,"F":12,"G":11,"H":11,"I":10,"J":13,"K":11,"L":11,"M":13,"N":13,"O":11,"P":34})
title(ws,"DAILY LOG",
 "The only sheet anyone types numbers into. Filled every morning for the day before, by 09:00. Every other sheet reads from here.")
r=5
r=band(ws,r,"Type into the yellow cells only. Grey columns are calculated.",15)
HEAD1=r
for cols,lab,fill in [((5,8),"GE1: PAID MEDIA",F_BLUE),((9,12),"GE2: SCHOOL OUTREACH",F_ORANGE),((13,16),"CALCULATED",F_GREY)]:
    c=ws.cell(row=r,column=cols[0],value=lab); c.font=NAVY_B; c.alignment=CTR
    for i in range(cols[0],cols[1]+1): ws.cell(row=r,column=i).fill=fill
    ws.merge_cells(start_row=r,start_column=cols[0],end_row=r,end_column=cols[1])
r+=1
HDR=r
r=header(ws,r,["Date","Day","Spend (R)","Clicks","Signups","Sales","Schools visited","Calls","Signups captured","Sales",
               "Total sales","Cumulative","Target cum.","Variance","Note"])
FIRSTD=r
for i,d in enumerate(DAYS):
    rr=r+i
    c=ws.cell(row=rr,column=2,value=d); c.number_format=DAT; c.font=BODY_B
    ws.cell(row=rr,column=3,value=i+1).font=MUTED; ws.cell(row=rr,column=3).alignment=CTR
    for col,fmt in [(4,CUR),(5,NUM),(6,NUM),(7,NUM),(8,NUM),(9,NUM),(10,NUM),(11,NUM)]:
        cc=ws.cell(row=rr,column=col); cc.fill=F_YELLOW; cc.font=INPUT; cc.number_format=fmt; cc.alignment=RGT
    m=ws.cell(row=rr,column=12,value="=G%d+K%d"%(rr,rr)); m.number_format=NUM; m.font=BODY_B; m.fill=F_GREY
    n=ws.cell(row=rr,column=13,value="=SUM($L$%d:L%d)"%(FIRSTD,rr)); n.number_format=NUM; n.font=BODY_B; n.fill=F_GREY
    o=ws.cell(row=rr,column=14,value="=ROUND(%s/%s*C%d,0)"%(A["qt"],A["days"],rr)); o.number_format=NUM; o.font=MUTED; o.fill=F_GREY
    p=ws.cell(row=rr,column=15,value="=M%d-N%d"%(rr,rr)); p.number_format='+#,##0;-#,##0;0'; p.font=BODY_B; p.fill=F_GREY
    q=ws.cell(row=rr,column=16); q.fill=F_YELLOW; q.font=INPUT
LASTD=FIRSTD+13
r=LASTD+1
ws.cell(row=r,column=2,value="TOTAL").font=BAND; ws.cell(row=r,column=2).fill=F_NAVY
for i in range(2,17): ws.cell(row=r,column=i).fill=F_NAVY
for col,fmt in [(4,CUR),(5,NUM),(6,NUM),(7,NUM),(8,NUM),(9,NUM),(10,NUM),(11,NUM),(12,NUM)]:
    c=ws.cell(row=r,column=col,value="=SUM(%s%d:%s%d)"%(chr(64+col),FIRSTD,chr(64+col),LASTD))
    c.font=Font(name="Arial",sz=10,b=True,color="FFFFFFFF"); c.number_format=fmt; c.alignment=RGT
TOT=r
c=ws.cell(row=r,column=14,value="=%s"%A["qt"]); c.font=Font(name="Arial",sz=10,b=True,color="FFFFFFFF"); c.number_format=NUM; c.alignment=RGT
c=ws.cell(row=r,column=15,value="=L%d-N%d"%(TOT,TOT)); c.font=Font(name="Arial",sz=10,b=True,color="FFFBBF24"); c.number_format='+#,##0;-#,##0;0'; c.alignment=RGT
r+=2
ws.cell(row=r,column=2,value="EXAMPLE ROW, do not type over the log above. Delete this block once the team has filled day 1.").font=WARN
r+=1
ex=["Mon 24 Aug","1","R2 400","310","44","6","2","9","31","3","9","9","18","-9","Meta set A pulled at 14:00, CTR 0.6%"]
for i,v in enumerate(ex):
    c=ws.cell(row=r,column=2+i,value=v); c.font=MUTED_I; c.fill=F_CREAM; c.alignment=RGT if i>1 else TOP
r+=2
for t in ["Spend is what the platform reports as spent that day, not what was budgeted.",
          "Sales are taken from the reconciled Peach export, split by the UTM the buyer arrived through. A sale with no UTM goes in neither engine column and is noted.",
          "Signups captured (GE2) are parent sign ups taken at a visit, from the QR code or the paper sheet, entered the same day.",
          "Target cum. divides the qualifying target evenly across 14 days. It is a pace line, not a forecast."]:
    ws.cell(row=r,column=2,value=t).font=MUTED_I; r+=1
ws.freeze_panes="D%d"%FIRSTD

json.dump({"FIRSTD":FIRSTD,"LASTD":LASTD,"TOT":TOT,"PFIRST":FIRST,"PLAST":LAST},open(B+"rows.json","w"))
wb.save(B+"_p3.xlsx"); print("p3 ok", FIRSTD, LASTD, TOT)
