import datetime as dt, json
from openpyxl import load_workbook
from style import *
B="/tmp/claude-0/-home-user-bridgeapp-os/6418b099-52bd-5eb3-aab2-a68e43f06251/scratchpad/build/"
wb=load_workbook(B+"_p3.xlsx")
A=json.load(open(B+"addr.json")); RW=json.load(open(B+"rows.json"))
LOG="'05_Daily_Log'"; F,LD=RW["FIRSTD"],RW["LASTD"]
START=dt.date(2026,8,24); DAYS=[START+dt.timedelta(days=i) for i in range(14)]
def S(col): return "SUM(%s!%s%d:%s%d)"%(LOG,col,F,col,LD)

def metric(ws,r,lab,formula,fmt,src,good=None):
    ws.cell(row=r,column=2,value=lab).font=BODY
    c=ws.cell(row=r,column=3,value=formula); c.number_format=fmt; c.font=BODY_B; c.alignment=RGT; c.fill=F_GREY
    if good is not None:
        t=ws.cell(row=r,column=4,value=good); t.number_format=fmt; t.font=MUTED; t.alignment=RGT
    ws.cell(row=r,column=5,value=src).font=MUTED_I
    return r+1

# ------------------------------------------------------- 06_GE1_Paid_Media
ws=wb.create_sheet("06_GE1_Paid_Media")
widths(ws,{"B":38,"C":16,"D":14,"E":56})
title(ws,"GE1: PAID MEDIA","Owner: Performance Marketer. Every figure reads from 05_Daily_Log. Nothing on this sheet is typed except the channel split below.")
r=5
r=band(ws,r,"Sprint to date",4); r=header(ws,r,["Metric","Actual","Target","Note"])
r=metric(ws,r,"Spend","=%s"%S("D"),CUR,"Sum of the log.","=%s"%A["cap"])
r=metric(ws,r,"Clicks","=%s"%S("E"),NUM,"Sum of the log.")
r=metric(ws,r,"Signups","=%s"%S("F"),NUM,"Sum of the log.")
r=metric(ws,r,"Sales","=%s"%S("G"),NUM,"Sum of the log. Peach reconciled, UTM attributed.","=%s"%A["t1"])
r=metric(ws,r,"Cost per sale","=IFERROR(%s/%s,0)"%(S("D"),S("G")),CUR,"The number this sprint exists to establish.","=%s"%A["cpa"])
r=metric(ws,r,"Cost per signup","=IFERROR(%s/%s,0)"%(S("D"),S("F")),CUR,"v4 GE1 KPI target is R85, current R110.",110)
r=metric(ws,r,"Click to signup rate","=IFERROR(%s/%s,0)"%(S("F"),S("E")),PCT,"Landing page performance.")
r=metric(ws,r,"Signup to sale rate","=IFERROR(%s/%s,0)"%(S("G"),S("F")),PCT,"v4 GE1 KPI target is 4%.",0.04)
r=metric(ws,r,"Spend left in the cap","=%s-%s"%(A["cap"],S("D")),CUR,"Negative means the cap is breached. That needs MD approval.")
r=metric(ws,r,"Share of GE1 target banked","=IFERROR(%s/%s,0)"%(S("G"),A["t1"]),PCT,"Against the GE1 split of the qualifying target.",1)
r+=1
r=band(ws,r,"Channel split. Type into the yellow cells",4)
r=header(ws,r,["Channel","Budget (R)","Spend (R)","Sales / cost per sale"])
CH=[("Meta, paid social",None),("Google, paid search",None),("LinkedIn, paid",None),("Boosted organic posts",None),("Influencer Wave 3, GP creator",50000)]
cf=r
for name,pref in CH:
    ws.cell(row=r,column=2,value=name).font=BODY
    c=ws.cell(row=r,column=3,value=pref); c.fill=F_YELLOW; c.font=INPUT; c.number_format=CUR; c.alignment=RGT
    c=ws.cell(row=r,column=4); c.fill=F_YELLOW; c.font=INPUT; c.number_format=CUR; c.alignment=RGT
    c=ws.cell(row=r,column=5); c.fill=F_YELLOW; c.font=INPUT; c.number_format=NUM; c.alignment=RGT
    r+=1
cl=r-1
ws.cell(row=r,column=2,value="Channel total").font=BODY_B
for col in (3,4,5):
    c=ws.cell(row=r,column=col,value="=SUM(%s%d:%s%d)"%(chr(64+col),cf,chr(64+col),cl))
    c.font=BODY_B; c.number_format=CUR if col<5 else NUM; c.alignment=RGT; c.fill=F_GREY
r+=1
ws.cell(row=r,column=2,value="Column E is sales, not cost per sale. Cost per sale per channel is D divided by E and is read at the Monday call.").font=MUTED_I
r+=2
ws.cell(row=r,column=2,value="Influencer Wave 3 budget R50 000 is the v4 Budget Master line for Aug 2026. Confirm how much of it falls inside these 14 days.").font=MUTED_I

# -------------------------------------------------- 07_GE2_School_Outreach
ws=wb.create_sheet("07_GE2_School_Outreach")
widths(ws,{"B":40,"C":16,"D":14,"E":60})
title(ws,"GE2: SCHOOL OUTREACH",
 "Owner: Schools Coordinator. A school visit produces four things before it produces a sale, so all four are tracked here. Reads from 05_Daily_Log.")
r=5
r=band(ws,r,"Read this before wondering why sales say zero",4)
for a,b in [
 ("Different windows","The Pretoria school outreach sprint ran 17 to 28 August. This workbook covers 24 August to 6 September. Week 1 of the outreach sits in the outreach report, not in this log, so entering it here will not move these numbers."),
 ("Sales lag the work","Registrations is the only column that counts as a sale, and it stays at zero until a presentation has actually been delivered to learners. Two weeks of visits producing zero sales is the expected shape, not a failure."),
 ("What moves instead","Schools contacted, meetings secured, learners reached and leads captured all move the moment they are logged. That is where the week's work shows up."),
]:
    ws.cell(row=r,column=2,value=a).font=BODY_B
    c=ws.cell(row=r,column=3,value=b); c.font=BODY; c.alignment=WRAP
    ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=5); ws.row_dimensions[r].height=28
    r+=1
r+=1
r=band(ws,r,"The pipeline, 24 August to 6 September",4); r=header(ws,r,["Stage","Actual","Target","Note"])
r=metric(ws,r,"Schools contacted","=%s"%S("H"),NUM,"Visits, calls and emails that reached someone.")
r=metric(ws,r,"Meetings or presentations secured","=%s"%S("I"),NUM,"A booked date, not a maybe.")
r=metric(ws,r,"Learners reached","=%s"%S("J"),NUM,"Only counts once a presentation has actually been delivered.")
r=metric(ws,r,"Leads captured","=%s"%S("K"),NUM,"Scanned the QR or signed the sheet.")
r=metric(ws,r,"Registrations","=%s"%S("L"),NUM,"The only stage that is a sale.","=%s"%A["t2"])
r+=1
r=band(ws,r,"Conversion, once there is something to convert",4); r=header(ws,r,["Rate","Actual","Target","Note"])
r=metric(ws,r,"Meetings per school contacted","=IFERROR(%s/%s,0)"%(S("I"),S("H")),PCT,"How often a contact turns into a booked date.")
r=metric(ws,r,"Learners per presentation","=IFERROR(%s/%s,0)"%(S("J"),S("I")),'0.0',"Room size. Tells you what a presentation is worth.")
r=metric(ws,r,"Leads per learner reached","=IFERROR(%s/%s,0)"%(S("K"),S("J")),PCT,"How well the QR and the pitch work in the room.")
r=metric(ws,r,"Lead to registration rate","=IFERROR(%s/%s,0)"%(S("L"),S("K")),PCT,"The number that decides whether schools can carry a sales target.")
r=metric(ws,r,"Registrations per school contacted","=IFERROR(%s/%s,0)"%(S("L"),S("H")),'0.00',"The school engine equivalent of cost per sale.")
r+=1
r=band(ws,r,"Baseline from the outreach report, 17 to 28 August",4)
r=header(ws,r,["Measure","Count","","Source"])
SRC="GradesMatch and UniApply School Outreach Report, Pretoria, internal, 29 August 2026."
for lab,v in [("Schools visited or contacted",27),("Schools requiring follow-up",7),
              ("Schools with active opportunities",5),("Access denied without an appointment",3),
              ("Confirmed Grade 11 presentation dates",2),("Learners reached",0),("UniApply registrations",0)]:
    ws.cell(row=r,column=2,value=lab).font=BODY
    c=ws.cell(row=r,column=3,value=v); c.font=BODY_B; c.number_format=NUM; c.alignment=RGT; c.fill=F_GREY
    ws.cell(row=r,column=5,value=SRC if lab.startswith("Schools visited") else "").font=MUTED_I
    r+=1
r+=1
r=band(ws,r,"What this engine costs",4)
r=header(ws,r,["Cost line","Sprint cost (R)","","Source"])
gf=r
for name,val,src in [
 ("Uber voucher, allocated",2000,"Allocated for school-to-school travel. Outreach report."),
 ("Uber voucher, spent to date","","Update after each travel day."),
 ("Acudeo Family Fun Day stall","","R1 000 if approved. Not covered by the voucher. Decision needed."),
 ("Printed leave behinds and parent slips",0,"Printed in house."),
]:
    ws.cell(row=r,column=2,value=name).font=BODY
    c=ws.cell(row=r,column=3,value=val)
    if val=="" : c.fill=F_YELLOW; c.font=INPUT
    else: c.font=BODY_B; c.fill=F_GREY
    c.number_format=CUR; c.alignment=RGT
    ws.cell(row=r,column=5,value=src).font=MUTED_I
    r+=1
gl=r-1
ws.cell(row=r,column=2,value="Cost per registration, school engine").font=BODY_B
c=ws.cell(row=r,column=3,value="=IFERROR(C%d/%s,0)"%(gf+1,S("L"))); c.font=BODY_B; c.number_format=CUR; c.alignment=RGT; c.fill=F_GREY
t=ws.cell(row=r,column=4,value="=%s"%A["cpa"]); t.number_format=CUR; t.font=MUTED; t.alignment=RGT
ws.cell(row=r,column=5,value="Voucher spent divided by registrations. Directly comparable with the paid media cost per sale on 06.").font=MUTED_I

wb.save(B+"_p4.xlsx"); print("p4 ok")
