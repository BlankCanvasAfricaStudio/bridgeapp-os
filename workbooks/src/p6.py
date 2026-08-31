import json
from openpyxl import load_workbook
from style import *
B="/tmp/claude-0/-home-user-bridgeapp-os/6418b099-52bd-5eb3-aab2-a68e43f06251/scratchpad/build/"
wb=load_workbook(B+"_p5.xlsx"); A=json.load(open(B+"addr.json"))
RW=json.load(open(B+"rows.json")); R2=json.load(open(B+"rows2.json"))
LOG="'05_Daily_Log'"; G1="'06_GE1_Paid_Media'"; G2="'07_GE2_School_Outreach'"; UTM="'09_UTM_Register'"
F,LD,TOT=RW["FIRSTD"],RW["LASTD"],RW["TOT"]
g1=lambda n:"%s!$C$%d"%(G1,n); g2=lambda n:"%s!$C$%d"%(G2,n)
DAYSLOG="COUNT(%s!$D$%d:$D$%d)"%(LOG,F,LD)
TAGGED="COUNTIF(%s!$H$%d:$H$%d,\"Y\")"%(UTM,R2["UF"],R2["UL"])
NREG=R2["UL"]-R2["UF"]+1
TOTSALES="%s!$L$%d"%(LOG,TOT)
TOTSPEND="%s!$D$%d"%(LOG,TOT)
GE2COST=g2(23)

# ----------------------------------------------------------------- 10_KPIs
ws=wb.create_sheet("10_KPIs")
widths(ws,{"B":40,"C":7,"D":10,"E":22,"F":14,"G":14,"H":14,"I":12,"J":12})
title(ws,"KPI DASHBOARD","Current reads live from 05_Daily_Log. Target and Trigger are fixed before the sprint starts and are not edited during it. Read at every Monday stand up.")
r=5
def grp(r,label,fill):
    return band(ws,r,label,8)
def kpi(r,name,cad,owner,cur,tgt,trg,good,fmt):
    ws.cell(row=r,column=2,value=name).font=BODY
    ws.cell(row=r,column=3,value=GE).font=BODY_B; ws.cell(row=r,column=3).alignment=CTR
    ws.cell(row=r,column=4,value=cad).font=MUTED; ws.cell(row=r,column=4).alignment=CTR
    ws.cell(row=r,column=5,value=owner).font=MUTED
    c=ws.cell(row=r,column=6,value=cur); c.number_format=fmt; c.font=BODY_B; c.alignment=RGT; c.fill=F_GREY
    c=ws.cell(row=r,column=7,value=tgt); c.number_format=fmt; c.font=NAVY_B; c.alignment=RGT
    c=ws.cell(row=r,column=8,value=trg); c.number_format=fmt; c.font=WARN; c.alignment=RGT
    ws.cell(row=r,column=9,value=good).font=MUTED; ws.cell(row=r,column=9).alignment=CTR
    if good in ("higher","lower"):
        cmp='F%d<H%d'%(r,r) if good=="higher" else 'F%d>H%d'%(r,r)
        core='IF(%s,"REVIEW","OK")'%cmp
        if good=="lower": core='IF(F%d=0,"NO DATA",%s)'%(r,core)
        f='=IF(NOT(ISNUMBER(H%d)),"MANUAL",%s)'%(r,core)
    else:
        f="MANUAL"
    c=ws.cell(row=r,column=10,value=f); c.font=BODY_B; c.alignment=CTR; c.fill=F_GREY
    return r+1

GE="GE1"
r=grp(r,"GE1: Paid media. Owner: Performance Marketer",F_BLUE)
r=header(ws,r,["KPI","GE","Cadence","Owner","Current","Target","Trigger","Good when","Status"])
r=kpi(r,"Cost per sale, blended paid","W","Performance Marketer","=%s"%g1(11),"=%s"%A["cpa"],"=%s*2"%A["cpa"],"lower",CUR)
r=kpi(r,"Sales banked, GE1","W","Performance Marketer","=%s"%g1(10),"=%s"%A["t1"],"=ROUND(%s*0.5,0)"%A["t1"],"higher",NUM)
r=kpi(r,"Cost per signup","W","Performance Marketer","=%s"%g1(12),85,120,"lower",CUR)
r=kpi(r,"Signup to sale rate","W","Performance Marketer","=%s"%g1(14),0.04,0.025,"higher",PCT)
r=kpi(r,"Spend left in the cap","W","Performance Marketer","=%s"%g1(15),0,0,"higher",CUR)
r+=1
GE="GE2"
r=grp(r,"GE2: School outreach. Owner: Schools Coordinator",F_ORANGE)
r=header(ws,r,["KPI","GE","Cadence","Owner","Current","Target","Trigger","Good when","Status"])
r=kpi(r,"Sales banked, GE2","W","Schools Coordinator","=%s"%g2(11),"=%s"%A["t2"],"=ROUND(%s*0.5,0)"%A["t2"],"higher",NUM)
r=kpi(r,"Cost per sale, school engine","W","Schools Coordinator","=%s"%g2(24),"=%s"%A["cpa"],"=%s*2"%A["cpa"],"lower",CUR)
r=kpi(r,"Schools visited","W","Schools Coordinator","=%s"%g2(7),12,8,"higher",NUM)
r=kpi(r,"Calls plus visits","W","Schools Coordinator","=%s"%g2(9),50,30,"higher",NUM)
r=kpi(r,"Signups captured","W","Schools Coordinator","=%s"%g2(10),"Baseline","Baseline","higher",NUM)
r=kpi(r,"Signups per visit","W","Schools Coordinator","=%s"%g2(12),"Baseline","Baseline","higher",'0.0')
r+=1
GE="ALL"
r=grp(r,"Sprint discipline. Owner: Campaign Lead",F_CREAM)
r=header(ws,r,["KPI","GE","Cadence","Owner","Current","Target","Trigger","Good when","Status"])
r=kpi(r,"Days logged in 05_Daily_Log","D","Campaign Lead","=%s"%DAYSLOG,14,12,"higher",NUM)
r=kpi(r,"Sprint codes created in the live UTM register","W","Performance Marketer","=%s"%TAGGED,NREG,NREG,"higher",NUM)
r=kpi(r,"Peach feed landing daily","D","Gradesmatch tech","Pass","Pass","Any miss","manual",'General')
r=kpi(r,"Visit logged the same day","D","Schools Coordinator","Pass","Pass","Any miss","manual",'General')
r=kpi(r,"Working book updated Friday","W","Campaign Lead","Pass","Pass","Any miss","manual",'General')
r+=1
ws.cell(row=r,column=2,value="Baseline means there is no honest number to set yet. These are read for four weeks and the target is set off real data.").font=MUTED_I
r+=1
ws.cell(row=r,column=2,value="MANUAL in the Status column means the row is judged by a person, not by a comparison. Cost per signup R85 target and R120 trigger, and the 4% signup to sale target, are from v4 sheet 10_KPIs_By_GE.").font=MUTED_I

# ------------------------------------------------------------ 11_Go_No_Go
ws=wb.create_sheet("11_Go_No_Go")
widths(ws,{"B":48,"C":18,"D":18,"E":12,"F":58})
title(ws,"DAY 14 GO OR NO GO","Completed on Sun 6 Sep 2026 by the Campaign Lead, decided by the MD. The thresholds below are set before the sprint starts and are not moved afterwards.")
r=5
r=band(ws,r,"The seven tests",5)
r=header(ws,r,["Test","Result","Threshold","Verdict","Note"])
TF=r
def test(r,name,res,thr,fmt,good,note,zero_is_nodata=False):
    ws.cell(row=r,column=2,value=name).font=BODY
    c=ws.cell(row=r,column=3,value=res); c.number_format=fmt; c.font=BODY_B; c.alignment=RGT; c.fill=F_GREY
    c=ws.cell(row=r,column=4,value=thr); c.number_format=fmt; c.font=NAVY_B; c.alignment=RGT
    cmp='C%d>=D%d'%(r,r) if good=="higher" else 'C%d<=D%d'%(r,r)
    core='IF(%s,"PASS","FAIL")'%cmp
    if zero_is_nodata: core='IF(C%d=0,"NO DATA",%s)'%(r,core)
    c=ws.cell(row=r,column=5,value='=IF(NOT(ISNUMBER(C%d)),"NO DATA",%s)'%(r,core))
    c.font=BODY_B; c.alignment=CTR; c.fill=F_GREY
    ws.cell(row=r,column=6,value=note).font=MUTED_I
    return r+1
BLENDCPS="IFERROR((%s+%s)/%s,0)"%(TOTSPEND,GE2COST,TOTSALES)
r=test(r,"1. Qualifying sales achieved","=%s"%TOTSALES,"=%s"%A["qt"],NUM,"higher","Total banked across both engines.")
r=test(r,"2. Blended cost per sale at or under target","=%s"%BLENDCPS,"=%s"%A["cpa"],CUR,"lower","Paid media spend plus school engine cost, divided by all sales.",zero_is_nodata=True)
r=test(r,"3. Paid engine at or under twice target","=%s"%g1(11),"=%s*2"%A["cpa"],CUR,"lower","A paid engine above this is not scalable at the current creative.",zero_is_nodata=True)
r=test(r,"4. School engine at or under twice target","=%s"%g2(24),"=%s*2"%A["cpa"],CUR,"lower","Same test, applied to visits.",zero_is_nodata=True)
r=test(r,"5. At least one engine at or under target",'=IF(OR(AND(ISNUMBER(%s),%s<=%s,%s>0),AND(ISNUMBER(%s),%s<=%s,%s>0)),1,0)'%(
        g1(11),g1(11),A["cpa"],g1(11),g2(24),g2(24),A["cpa"],g2(24)),1,NUM,"higher",
        "One working engine is enough to justify releasing budget into it.")
r=test(r,"6. Paid spend stayed inside the cap","=%s"%TOTSPEND,"=%s"%A["cap"],CUR,"lower","A breach needs MD sign off recorded below.",zero_is_nodata=True)
r=test(r,"7. Days logged","=%s"%DAYSLOG,14,NUM,"higher","A sprint that was not logged did not measure anything, whatever the sales say.")
TL=r-1
r+=1
c=ws.cell(row=r,column=2,value="Tests passed"); c.font=BODY_B
c=ws.cell(row=r,column=3,value='=COUNTIF(E%d:E%d,"PASS")'%(TF,TL)); c.font=BODY_B; c.number_format=NUM; c.alignment=RGT; c.fill=F_GREY
ws.cell(row=r,column=4,value=7).font=NAVY_B; ws.cell(row=r,column=4).alignment=RGT
ws.cell(row=r,column=6,value="NO DATA means the test had nothing to judge. Before the sprint runs, most tests read NO DATA. That is correct.").font=MUTED_I
r+=1
ws.cell(row=r,column=2,value="Tests still reading NO DATA").font=BODY_B
c=ws.cell(row=r,column=3,value='=COUNTIF(E%d:E%d,"NO DATA")'%(TF,TL)); c.font=BODY_B; c.number_format=NUM; c.alignment=RGT; c.fill=F_GREY
ws.cell(row=r,column=6,value="Must be zero on day 14. A test with no data is a measurement failure, not a neutral result.").font=MUTED_I
r+=2
r=band(ws,r,"What the result costs to scale",5)
r=header(ws,r,["Measure","Value","","","Note"])
rem="(%s-%s)"%(A["full"],TOTSALES)
for lab,f,fmt,note in [
 ("Sales still needed after this sprint","=%s"%rem,NUM,"Against the 40 000 target."),
 ("Cost to buy them at the blended rate achieved","=%s*%s"%(rem,BLENDCPS),CUR,"This is the number the MD is actually deciding on."),
 ("Cost to buy them at the target rate","=%s*%s"%(rem,A["cpa"]),CUR,"What the plan assumed."),
 ("Difference","=%s*%s-%s*%s"%(rem,BLENDCPS,rem,A["cpa"]),CUR,"Positive means the campaign costs more than planned."),
 ("Weeks left after this sprint","=%s-2"%A["wks"],NUM,"From 7 Sep to 30 Nov 2026."),
 ("Sales per week required from here","=IFERROR(ROUND(%s/(%s-2),0),0)"%(rem,A["wks"]),NUM,"Compare against what the sprint actually produced in two weeks."),
]:
    ws.cell(row=r,column=2,value=lab).font=BODY
    c=ws.cell(row=r,column=3,value=f); c.number_format=fmt; c.font=BODY_B; c.alignment=RGT; c.fill=F_GREY
    ws.cell(row=r,column=6,value=note).font=MUTED_I
    r+=1
r+=1
r=band(ws,r,"The decision",5)
for lab,note in [("Decision","RELEASE FULL BUDGET / RELEASE INTO ONE ENGINE ONLY / HOLD AND RE-TEST / RE-CUT THE TARGET"),
                 ("Reason","One or two sentences. What the evidence showed."),
                 ("Budget released (R)","Only if the decision is a release."),
                 ("Decided by",""),("Date",""),("Cap breach approved by","Leave blank if test 6 passed.")]:
    ws.cell(row=r,column=2,value=lab).font=BODY_B
    c=ws.cell(row=r,column=3); c.fill=F_YELLOW; c.font=INPUT
    ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=5)
    ws.cell(row=r,column=6,value=note).font=MUTED_I
    ws.row_dimensions[r].height=18
    r+=1
wb.save(B+"_p6.xlsx"); print("p6 ok")
