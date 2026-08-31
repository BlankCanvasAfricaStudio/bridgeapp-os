import datetime as dt, json
from openpyxl import load_workbook
from style import *
B="/tmp/claude-0/-home-user-bridgeapp-os/6418b099-52bd-5eb3-aab2-a68e43f06251/scratchpad/build/"
wb=load_workbook(B+"_p1b.xlsx"); K=json.load(open(B+"bas.json"))
DRV="'03_Sprint_Drivers'"
START=dt.date(2026,8,24); DAYS=[START+dt.timedelta(days=i) for i in range(14)]
A={}
ws=wb.create_sheet("03_Sprint_Drivers")
widths(ws,{"B":44,"C":16,"D":13,"E":62})
title(ws,"SPRINT DRIVERS",
 "Every assumption in one cell. Yellow cells are set by a person. Everything else calculates. Cost per sale is derived here, never assumed.")
r=5
def drv(row,label,val,fmt,src,key=None,formula=False,fill=True):
    ws.cell(row=row,column=2,value=label).font=BODY
    c=ws.cell(row=row,column=3,value=val)
    if formula: c.font=BODY_B; c.fill=F_GREY
    else:
        c.font=INPUT
        if fill: c.fill=F_YELLOW
    c.number_format=fmt; c.alignment=RGT
    ws.cell(row=row,column=5,value=src).font=MUTED_I
    if key: A[key]="%s!$C$%d"%(DRV,row)
    return row+1

r=band(ws,r,"Sprint window",4); r=header(ws,r,["Driver","Value","","Source or note"])
r=drv(r,"Sprint start",START,DAT,"Monday. First full week after the 21 Aug data pull.","start")
r=drv(r,"Sprint end",DAYS[-1],DAT,"Sunday. Two complete weeks.","end")
r=drv(r,"Days in sprint","=%s-%s+1"%(A["end"],A["start"]),NUM,"Calculated.","days",formula=True)
r+=1

r=band(ws,r,"What is being risked, and what it has to prove",4)
r=header(ws,r,["Driver","Value","","Source or note"])
r=drv(r,"Sprint media cap, GE1 paid (R)",10000,CUR,
 "Agreed 21 Aug: R5 000 a week across both weeks. Sits inside the R250 000 of paid media the v4 plan already carries rather than adding to it.","cap")
r=drv(r,"Qualifying sales target",61,NUM,
 "Agreed 21 Aug. What R10 000 returns if click to sale is 2%, which is the hypothesis being tested. Hitting it confirms 2% and tells exco R2m buys about 12 100 sales, not 40 000.","qt")
r=drv(r,"Share of the target from GE1 paid",1.0,PCT,
 "100% this sprint. The 61 is what the paid budget buys. School outreach carries no media spend and is setting its first baseline, so it is measured rather than targeted.","sp1")
r=drv(r,"Share of the target from GE2 schools","=1-%s"%A["sp1"],PCT,"Calculated. Zero by design this sprint. GE2 sales are counted and reported, just not targeted.","sp2",formula=True)
r=drv(r,"GE1 sales target","=ROUND(%s*%s,0)"%(A["qt"],A["sp1"]),NUM,"Calculated.","t1",formula=True)
r=drv(r,"GE2 sales target","=%s-%s"%(A["qt"],A["t1"]),NUM,"Calculated. Remainder, so the two always sum to the target.","t2",formula=True)
r=drv(r,"Sales required per day","=ROUND(%s/%s,1)"%(A["qt"],A["days"]),'0.0',"Calculated.","pd",formula=True)
r+=1

r=band(ws,r,"What the real cost per click means for that target",4)
r=header(ws,r,["Driver","Value","","Source or note"])
r=drv(r,"Cost per link click","=%s"%K["cpc"],CUR2,
 "Real. From the July Meta boost on 02_Baselines. Not an estimate.","cpc",formula=True)
r=drv(r,"Clicks the cap buys at that rate","=ROUND(%s/%s,0)"%(A["cap"],A["cpc"]),NUM,"Calculated.","clk",formula=True)
r=drv(r,"Click to sale rate GE1 must achieve","=IFERROR(%s/%s,0)"%(A["t1"],A["clk"]),PCT,
 "Calculated. THE TEST. Nobody knows the real rate, because no click has ever been tied to a payment. For R2m to buy 40 000 this would have to reach 6.6% at today's cost per click.","req",formula=True)
r=drv(r,"Cost per sale this implies, GE1","=IFERROR(%s/%s,0)"%(A["cap"],A["t1"]),CUR,
 "Calculated from the cap and the target, not assumed. v4 sheet 09 sets R85 as the cost per free signup target.","cpa",formula=True)
r+=1

r=band(ws,r,"What the cap buys at different click to sale rates",4)
r=header(ws,r,["If click to sale is","Cost per sale becomes","Sales the cap buys",""])
SENS=r
for rate in [0.005,0.01,0.02,0.04,0.08]:
    c=ws.cell(row=r,column=2,value=rate); c.number_format=PCT; c.font=NAVY_B; c.alignment=RGT
    c=ws.cell(row=r,column=3,value="=%s/B%d"%(A["cpc"],r)); c.number_format=CUR; c.font=BODY_B; c.alignment=RGT; c.fill=F_GREY
    c=ws.cell(row=r,column=4,value="=ROUND(%s*B%d,0)"%(A["clk"],r)); c.number_format=NUM; c.font=BODY_B; c.alignment=RGT; c.fill=F_GREY
    r+=1
ws.cell(row=r,column=2,value="Read this against the target above. It is the whole commercial question in five rows.").font=MUTED_I
r+=2

r=band(ws,r,"Against the full campaign",4); r=header(ws,r,["Driver","Value","","Source or note"])
r=drv(r,"Full campaign target",40000,NUM,"Campaign brief. Paid transactions by 30 Nov 2026.","full")
r=drv(r,"Weeks from sprint start to deadline",15,NUM,"24 Aug to 30 Nov 2026.","wks")
r=drv(r,"Required pace per week","=ROUND(%s/%s,0)"%(A["full"],A["wks"]),NUM,"Calculated.","pw",formula=True)
r=drv(r,"Required sales in a 14 day window","=%s*2"%A["pw"],NUM,"Calculated. What full pace looks like over this sprint.","p14",formula=True)
r=drv(r,"This sprint as a share of full pace","=IFERROR(%s/%s,0)"%(A["qt"],A["p14"]),PCT,
 "Calculated. How far the qualifying target sits below full pace.","shr",formula=True)
r+=1

r=band(ws,r,"How a sale is counted in this sprint",4)
for a,b in [
 ("Primary source","An automated feed of cleared Peach transactions, delivered by Gradesmatch tech. This is a tech function. Nobody on the growth team has capacity to reconcile by hand every day."),
 ("Attribution to engine","By the UTM on the link the buyer arrived through, registered in 09_UTM_Register. A sale with no UTM is logged as unattributed."),
 ("GA4","Cannot attribute a sale until the purchase event is installed. In this sprint GA4 is used for traffic and page progression only."),
 ("Deduplication","On transaction_id. A transaction_id already in the ledger is not counted twice."),
 ("Cut off","23:59 each day. A payment clearing after cut off is logged against the following day."),
]:
    ws.cell(row=r,column=2,value=a).font=BODY_B
    c=ws.cell(row=r,column=3,value=b); c.font=BODY; c.alignment=WRAP
    ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=5); ws.row_dimensions[r].height=26; r+=1
r+=1
r=band(ws,r,"Blocked on Gradesmatch dev at the time of writing",4)
r=header(ws,r,["Item","Needed by","","Effect on this sprint if it does not land"])
for a,b,c in [
 ("GA4 purchase event with value","Day 1, 24 Aug","Without this OR the transaction feed below, no sale can be recorded at all and the sprint measures nothing. With the feed but not this, sales are counted but cannot be split by channel."),
 ("Landing page live on bridgeapp.co.za","Day 1, 24 Aug","Paid traffic has no dedicated destination and converts lower. Change the URL on every row of 09_UTM_Register before spending."),
 ("Lead endpoint (window.BRIDGEAPP_LEAD_ENDPOINT)","Day 1, 24 Aug","Readiness tracker submissions are not captured, so click to signup stays unmeasured and only click to sale can be read."),
 ("Automated feed of cleared Peach transactions","Day 1, 24 Aug","The manual reconciliation this replaces is no longer resourced. Without this or the purchase event there is no way to count a sale."),
]:
    ws.cell(row=r,column=2,value=a).font=BODY; ws.cell(row=r,column=2).fill=F_ORANGE
    ws.cell(row=r,column=3,value=b).font=BODY_B; ws.cell(row=r,column=3).fill=F_ORANGE
    cc=ws.cell(row=r,column=5,value=c); cc.font=MUTED; cc.alignment=WRAP; cc.fill=F_ORANGE
    ws.row_dimensions[r].height=26; r+=1
json.dump(A,open(B+"addr.json","w"))
wb.save(B+"_p2.xlsx"); print("p2 ok"); print(json.dumps(A,indent=0))
